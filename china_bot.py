import discord
from discord.ext import commands, tasks
from discord import app_commands
import openpyxl
import os
import time
import random
import ast
import dotenv

intents = discord.Intents.default()
intents.message_content = True

bot = discord.Client(command_prefix="!", intents=intents)
tree = app_commands.CommandTree(bot)

dotenv.load_dotenv()

dc_token = os.getenv("DCTOKEN")

bad_word = ["在外面", "沒空", "原批", "不要", "補習", "原P", "原p"] #文字獄
banned_word = [] #文字獄
good_word = ["要不要"] #避免誤判用

owner_id = int(os.getenv("OWNERID"))
bot_id = int(os.getenv("BOTID"))
owner_name = os.getenv("OWNERNAME")

print("Owner id: ", owner_id)
print("Bot id: ", bot_id)

reply_time = 10 #seconds 設定被@的人需要於幾秒內回覆

bad_message_id_count = 50 #紀錄藐視人的人

tagged = {} #紀錄被@的人
tagged_channel_id = {} #紀錄被@的人所在頻道

current_path = os.path.dirname(__file__)

running = True #運作狀態

#每一段時間檢查是否有人被tag沒於時間內回覆
@tasks.loop(seconds=2)
async def check_response():
        global tagged
        global tagged_channel_id
        temp_tagged = tagged
        for tag in temp_tagged:
                if int(time.time()) - int(tagged[tag]) > reply_time:
                        wb = openpyxl.load_workbook(current_path + "/users.xlsx")
                        ws = wb.active
                        ws.title = "users"
                        user = await bot.fetch_user(tag)
                        print(user.name)
                        ws.append([f"壞寶寶:{str(user.name)}沒在{reply_time}秒內回{owner_name}的訊息"])
                        wb.save(current_path + "/users.xlsx")
                        user = await bot.fetch_user(tag)
                        print(f"已新增[{user.name}]至名單")
                        print(int(time.time()) - int(tagged[tag]))
                        channel = await bot.fetch_channel(tagged_channel_id[tag])
                        file = discord.File("bruh.mp3")
                        await channel.send(f"<@{tag}>拒絕答覆? 你違反了藐視{owner_name}罪，太離譜了 :rage:", file=file)
                        tagged.pop(tag)
                        tagged_channel_id.pop(tag)
                        return
        if not running:
                check_response.stop()

#接收訊息
@bot.event
async def on_message(message: discord.Message):
        global tagged
        print(f"從[{message.channel}]由[{message.author}]所發布，內容為:{message.content}")
        if not running:
                return #不在運作模式下
        did =False
        if str(message.author) != str(bot.user):
                if f"<@{bot_id}>" in str(message.content) or f"<@{owner_id}>" in str(message.content): #機器人或擁有者被tag
                        file = discord.File("bruh.mp3")
                        #紀錄有藐視的訊息
                        with open("bad_message_id.txt", 'r') as f:
                                d:list = ast.literal_eval(f.read())
                        print(d)
                        if len(d) < bad_message_id_count:
                                d.append(int(message.id))
                                print(d)
                                with open("bad_message_id.txt", 'w+') as f:
                                        f.write(str(d))
                        else:
                                d.pop(0)
                                d.append(int(message.id))
                                print(d)
                        with open("bad_message_id.txt", 'w+') as f:
                                f.write(str(d))
                        await message.channel.send(f"<@{message.author.id}>你這是在反質詢嗎? 你違反了藐視{owner_name}罪，太離譜了 :rage:", file=file)
                        return
                #排除有防誤判詞的訊息
                for keyword in good_word:
                        if str(keyword) in str(message.content):
                                return
                #言論審查
                for keyword in bad_word:
                        if str(keyword) in str(message.content):
                                if message.author.id == owner_id:
                                        continue
                                channel = bot.get_channel(message.channel.id)
                                if did == False:
                                        wb = openpyxl.load_workbook(current_path + "/users.xlsx")
                                        ws = wb.active
                                        ws.title = "users"
                                        print(message.author)
                                        ws.append([f"壞寶寶:{str(message.author)}"])
                                        ws.append([f"說了:{str(message.content)}"])
                                        wb.save(current_path + "/users.xlsx")
                                        print(f"已新增[{message.author}]至名單")
                                        #記錄有藐視的訊息
                                        with open("bad_message_id.txt", 'r') as f:
                                                d:list = ast.literal_eval(f.read())
                                        if len(d) < bad_message_id_count:
                                                d.append(int(message.id))
                                                print(d)
                                                with open("bad_message_id.txt", 'w+') as f:
                                                        f.write(str(d))
                                        else:
                                                d.pop(0)
                                                d.append(int(message.id))
                                                with open("bad_message_id.txt", 'w+') as f:
                                                        f.write(str(d))
                                        file = discord.File("bruh.mp3")
                                        await channel.send(f"<@{message.author.id}>虛偽陳述! 你違反了藐視{owner_name}罪，太離譜了 :rage:", file=file)
                                        did = True
                
                for keyword in banned_word:
                        if str(keyword) in str(message.content):
                                if message.author.id == owner_id:
                                        continue
                                channel = bot.get_channel(message.channel.id)
                                if did == False:
                                        wb = openpyxl.load_workbook(current_path + "/users.xlsx")
                                        ws = wb.active
                                        ws.title = "users"
                                        print(message.author)
                                        ws.append([f"壞寶寶:{str(message.author)}"])
                                        ws.append([f"說了:{str(message.content)}"])
                                        wb.save(current_path + "/users.xlsx")
                                        print(f"已新增[{message.author}]至名單")
                                        #記錄有藐視的訊息
                                        with open("bad_message_id.txt", 'r') as f:
                                                d:list = ast.literal_eval(f.read())
                                        if len(d) < bad_message_id_count:
                                                d.append(int(message.id))
                                                print(d)
                                                with open("bad_message_id.txt", 'w+') as f:
                                                        f.write(str(d))
                                        else:
                                                d.pop(0)
                                                d.append(int(message.id))
                                                with open("bad_message_id.txt", 'w+') as f:
                                                        f.write(str(d))
                                        file = discord.File("bruh.mp3")
                                        await channel.send(f"<@{message.author.id}>你藐視{owner_name}! 你違反了藐視{owner_name}罪，太離譜了 :rage:", file=file)
                                        did = True

                #擁有者tag人
                if message.author.id == owner_id:
                        if "<@" in str(message.content):
                                if "<@#" in str(message.content):
                                        return
                                for i in message.mentions:
                                        mention = i
                                        #紀錄被tag人的id和頻道id
                                        tagged[mention.id] = time.time()
                                        print(mention.id, " tagged")
                                        tagged_channel_id[mention.id] = message.channel.id
                #被tag的人傳訊息
                if message.author.id in tagged:
                        #回覆時間朝過reply_time
                        if int(time.time()) - int(tagged[message.author.id]) >= reply_time:
                                wb = openpyxl.load_workbook(current_path + "/users.xlsx")
                                ws = wb.active
                                ws.title = "users"
                                print(message.author.name)
                                ws.append([f"壞寶寶:{str(message.author.name)}沒在{reply_time}秒內回{owner_name}的訊息"])
                                print(int(time.time()) - int(tagged[message.author.id]))
                                wb.save(current_path + "/users.xlsx")
                                print(f"已新增[{message.author}]至名單")
                                file = discord.File("bruh.mp3")
                                await message.channel.send(f"<@{message.author.id}>拒絕答覆? 你違反了藐視{owner_name}罪，太離譜了 :rage:", file=file)
                                tagged.pop(message.author.id)
                        #於時間內回覆
                        else:
                                print(time.time() - int(tagged[message.author.id]))
                                print(message.author.id, "responded")
                                tagged.pop(message.author.id)

#機器人，啟動!
@bot.event
async def on_ready():
        print("Logging as {}".format(bot.user))
        try:
                synced = await tree.sync()
                print(f"Synced {len(synced)} command(s)")
                game = discord.Game("言論審查")
                await bot.change_presence(activity=game)
                check_response.start()
        except Exception as e:
                print(e)

#偵測刪除訊息
@bot.event
async def on_message_delete(message: discord.Message):
        print(f"Message {message.content} was deleted by {message.author}")
        #載入藐視的訊息
        with open("bad_message_id.txt", 'r') as f:
                d:list = ast.literal_eval(f.read())
        for id in d:
                if message.id == id:
                        #被刪除訊息為藐視的訊息 (試圖讓機器人尷尬)
                        await message.channel.send(str(message.content))
                        file = discord.File("bruh.mp3")
                        await message.channel.send(f"<@{message.author.id}>隱匿資訊? 你違反了藐視{owner_name}罪，太離譜了 :rage:", file=file)

#查看禁字指令
@tree.command(name="keyword_list", description="關鍵字列表")
async def keyword_list(ctx:discord.Interaction):
        print("{} did \'/keyword_list\'".format(ctx.user))
        text = []
        text = f"禁字: {bad_word}, {banned_word}\n允許字: {good_word}"
        await ctx.response.send_message(text)

#開關機器人運作模式指令
@tree.command(name="switch", description="開關")
async def switch(ctx:discord.Interaction):
        print("{} did \'/switch\'".format(ctx.user))
        if ctx.user.id != owner_id:
                #排除執行指令者非擁有者
                choose = random.randint(0, 1)
                if choose:
                        await ctx.response.send_message("你超可悲，就算外面有10萬個人也沒辦法阻止我 :)")
                else:
                        file = discord.File("hahaha.mp3")
                        await ctx.response.send_message("😐喔😏哼哼😀哈哈哈😃哈哈哈😃喝喝喝😄呵呵呵", file=file)
                return
        global running
        if running: #正在運作 -> 關閉
                game = discord.Game("睡覺")
                await bot.change_presence(activity=game)
                running = False
                await ctx.response.send_message("該去晉見謝主席ㄌ :dog:")
        else: #停止運作 -> 啟動
                game = discord.Game("言論審查")
                await bot.change_presence(activity=game)
                running = True
                check_response.start()
                await ctx.response.send_message("該開始言論審查ㄌ :smiling_imp:")

#主程式
if __name__ == "__main__":
        bot.run(dc_token)